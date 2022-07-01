from PyQt5.QtWidgets import (
    QApplication, QWidget,
    QVBoxLayout, QListWidget,
    QLineEdit, QPushButton,
    QHBoxLayout, QListWidgetItem,
    QFileDialog
)
from PyQt5.QtGui import QIcon, QFont, QGuiApplication
from PyQt5.QtCore import Qt, pyqtSignal
from src.constants import *
from src.newlink import NewLink
from db.manager import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class LinksSettings(QWidget):

    close_links = pyqtSignal()

    def __init__(self, db: Manager, parent=None):
        super(LinksSettings, self).__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        self.db = db
        self.setup_UI()
        self.destroyed.connect(LinksSettings._on_destroyed)

    @staticmethod
    def _on_destroyed():
        print("LinksSettings instance deleted.")

    def setup_UI(self):
        self.setWindowIcon(QIcon(":/hyperlink.png"))
        self.setWindowTitle("URL Links")
        self.mainlayout = QVBoxLayout()
        self.selectedlinklayout = QHBoxLayout()
        self.btnslayout = QHBoxLayout()
        self.setup_url_list()
        self.setup_btns()
        self.update_url_list()
        self.setLayout(self.mainlayout)
        self.setMinimumWidth(640)

    def setup_url_list(self):
        font = QFont()
        font.setPointSize(12)
        font.setFamily("Roboto Mono")
        self.urllist = QListWidget()
        self.urllist.itemClicked.connect(
            lambda: self.urllink.setText(
                get_app_link(self.db, name=self.urllist.currentItem().text().replace(" ", "_").lower()).url
            )
        )
        self.urllist.setFont(font)
        self.urllink = QLineEdit()
        self.urllink.setFont(font)
        self.selectedlinklayout.addWidget(self.urllink, 60)
        self.mainlayout.addWidget(self.urllist)
        self.mainlayout.addLayout(self.selectedlinklayout)

    def setup_btns(self):
        font = QFont()
        font.setPointSize(12)
        font.setFamily("Roboto Mono")
        self.btns_conf = {
            "copy": ("Copy", self.copy),
            "update": ("Update", self.update)
        }
        self.btns = {}
        self.btns["new"] = QPushButton("New")
        self.btns["new"].clicked.connect(self.open_add_new_link)
        self.btns["new"].setFont(font)
        self.btnslayout.addWidget(self.btns["new"])
        self.btns["export"] = QPushButton("Export")
        self.btns["export"].clicked.connect(self.export_links)
        self.btns["export"].setFont(font)
        self.btnslayout.addWidget(self.btns["export"])
        self.btns["delete"] = QPushButton("Delete")
        self.btns["delete"].clicked.connect(self.delete_link)
        self.btns["delete"].setFont(font)
        self.btnslayout.addWidget(self.btns["delete"])
        self.btnslayout.addStretch()
        for name, val in self.btns_conf.items():
            btn = QPushButton(val[0])
            btn.clicked.connect(val[1])
            btn.setFont(font)
            self.selectedlinklayout.addWidget(btn)
            self.btns[name] = btn
        self.mainlayout.addLayout(self.btnslayout)

    def closeEvent(self, event):
        self.close_links.emit()
        return super().closeEvent(event)

    def copy(self):
        clipboard = QGuiApplication.clipboard()
        clipboard.setText(self.urllink.text())

    def update(self):
        if self.urllist.currentItem() is not None:
            update_app_link(self.db, self.urllist.currentItem().text().replace(" ", "_").lower(), self.urllink.text())
            self.close_links.emit()
            self.close()

    def delete_link(self):
        if self.urllist.currentItem() is not None:
            self.urllink.clear()
            delete_app_link(self.db, self.urllist.currentItem().text().replace(" ", "_").lower())
            self.update_url_list()

    def update_url_list(self):
        self.urllist.clear()
        for link in get_app_links(self.db):
            self.urllist.addItem(QListWidgetItem(link.name.replace("_", " ").upper()))

    def open_add_new_link(self):
        self.newlink = NewLink(self.db)
        self.newlink.setWindowModality(Qt.WindowModality.ApplicationModal)
        self.newlink.open_link_settings.connect(self.show)
        self.newlink.update_url_list.connect(self.update_url_list)
        self.newlink.show()
        self.hide()

    def export_links(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Export links", filter="XLSX File (*.xlsx)")
        if len(filepath) > 0:
            links = get_all_links(self.db)
            wb = Workbook()
            ws = wb.active
            ws.title = "PCC URL LINKS"
            for link in links:
                ws.append([link.name, link.group, link.url])
            ws.column_dimensions[get_column_letter(1)].width = 40
            ws.column_dimensions[get_column_letter(2)].width = 40
            ws.column_dimensions[get_column_letter(3)].width = 200
            wb.save(filepath)